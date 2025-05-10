#! /usr/bin/python
#
# Description:
# ================================================================
# Time-stamp: "2025-03-13 01:29:19 junaid"
# ================================================================
#
# Author:  Muhammad Junaid III <mjo147@uregina.ca>
#
# Copyright (c) junaid
#
###################################################################################################################################################

# Import relevant packages
import uproot
import uproot as up
import numpy as np

np.bool = bool
np.float = float

import root_numpy as rnp
import pandas as pd
import root_pandas as rpd
import ROOT
import scipy
import scipy.integrate as integrate
import matplotlib.pyplot as plt
import sys, math, os, subprocess
import array
import csv
from ROOT import TCanvas, TList, TPaveLabel, TColor, TGaxis, TH1F, TH2F, TPad, TStyle, gStyle, gPad, TLegend, TGaxis, TLine, TMath, TLatex, TPaveText, TArc, TGraphPolar, TText, TString
from ROOT import kBlack, kCyan, kRed, kGreen, kMagenta, kBlue
from functools import reduce
import math as ma
import uncertainties as u
from ctypes import c_double

##################################################################################################################################################
ROOT.gROOT.SetBatch(ROOT.kTRUE) # Set ROOT to batch mode explicitly, does not splash anything to screen
###############################################################################################################################################

# Check the number of arguments provided to the script
if len(sys.argv)-1!=7:
    print("!!!!! ERROR !!!!!\n Expected 7 arguments\n Usage is with - ROOTfileSuffixs Beam Energy MaxEvents RunList CVSFile\n!!!!! ERROR !!!!!")
    sys.exit(1)

##################################################################################################################################################

# Input params - run number and max number of events
PHY_SETTING = sys.argv[1]
MaxEvent = sys.argv[2]
DATA_Suffix = sys.argv[3]
DUMMY_Suffix = sys.argv[4]
DATA_RUN_LIST = sys.argv[5]
DUMMY_RUN_LIST = sys.argv[6]
CSV_FILE = sys.argv[7]

################################################################################################################################################
'''
ltsep package import and pathing definitions
'''

# Import package for cuts
from ltsep import Root

lt=Root(os.path.realpath(__file__), "Plot_ProdCoin")

# Add this to all files for more dynamic pathing
USER=lt.USER # Grab user info for file finding
HOST=lt.HOST
REPLAYPATH=lt.REPLAYPATH
UTILPATH=lt.UTILPATH
ANATYPE=lt.ANATYPE
OUTPATH=lt.OUTPATH
RUNLISTPATH = "/u/group/c-pionlt/USERS/%s/hallc_replay_lt/UTIL_BATCH/InputRunLists/PionLT_2021_2022" % (USER)
EFF_CSV     = "/u/group/c-pionlt/USERS/%s/hallc_replay_lt/UTIL_PION/efficiencies" % (USER)
MMCUT_CSV   = "/u/group/c-pionlt/USERS/%s/hallc_replay_lt/UTIL_PION/LTSep_CSVs/mm_offset_cut_csv" % (USER)
DCUT_CSV    = "/u/group/c-pionlt/USERS/%s/hallc_replay_lt/UTIL_PION/LTSep_CSVs/diamond_cut_csv" % (USER)
TBINCSVPATH = "/u/group/c-pionlt/USERS/%s/hallc_replay_lt/UTIL_PION/LTSep_CSVs/t_binning_csv" % (USER)

#################################################################################################################################################

# Output PDF File Name
print("Running as %s on %s, hallc_replay_lt path assumed as %s" % (USER, HOST, REPLAYPATH))
#Pion_Analysis_Distributions = "%s/%s_%s_ProdCoin_Pion_Analysis_DataYield_Distributions.pdf" % (OUTPATH, PHY_SETTING, MaxEvent)

# Extract the first three words from PHY_SETTING for the CSV file name
setting_name = "_".join(PHY_SETTING.split("_")[:3])

# Input file location and variables taking
rootFile_DATA = "%s/%s_%s_%s.root" % (OUTPATH, PHY_SETTING, MaxEvent, DATA_Suffix)
rootFile_DUMMY = "%s/%s_%s_%s.root" % (OUTPATH, PHY_SETTING, MaxEvent, DUMMY_Suffix)
data_run_list = "%s/%s" % (RUNLISTPATH, DATA_RUN_LIST)
dummy_run_list = "%s/%s" % (RUNLISTPATH, DUMMY_RUN_LIST)
eff_csv_file = "%s/%s.csv" % (EFF_CSV, CSV_FILE)
mmcut_csv_file = "%s/%s_mm_offsets_cuts_parameters.csv" % (MMCUT_CSV, setting_name)
dcut_csv_file = "%s/%s_diamond_cut_parameters.csv" % (DCUT_CSV, setting_name)
tbin_csv_file  = "%s/%s_tbinning_yields_pions.csv" % (TBINCSVPATH, setting_name)

###################################################################################################################################################

print ('\nPhysics Setting = ',PHY_SETTING, '\n')
print("-"*40)

# Cuts for Pions Selection - Change according to your data
# Read the vertices from the CSV file
vertices = {}
try:
    with open(dcut_csv_file, mode='r') as csv_file:
        csv_reader = csv.DictReader(csv_file)
        for row in csv_reader:
            vertex_name = row["Vertex"]
            x_value = round(float(row["X"]), 3)  # Round to 3 decimal places
            y_value = round(float(row["Y"]), 3)  # Round to 3 decimal places
            vertices[vertex_name] = [x_value, y_value]
except (FileNotFoundError, ValueError, KeyError) as e:
    print(f"Error reading diamond cut vertices from {dcut_csv_file}: {e}")
    sys.exit(1)

# Assign the vertices
vertex1 = vertices["vertex1"]  # bottom-left
vertex2 = vertices["vertex2"]  # top-left
vertex3 = vertices["vertex3"]  # top-right
vertex4 = vertices["vertex4"]  # bottom-right

# Print the vertex values rounded to 3 decimals
#print("Diamond Cut Vertices (rounded to 3 decimals):")
#print(f"Vertex 1 (bottom-left): [{vertex1[0]:.3f}, {vertex1[1]:.3f}]")
#print(f"Vertex 2 (top-left): [{vertex2[0]:.3f}, {vertex2[1]:.3f}]")
#print(f"Vertex 3 (top-right): [{vertex3[0]:.3f}, {vertex3[1]:.3f}]")
#print(f"Vertex 4 (bottom-right): [{vertex4[0]:.3f}, {vertex4[1]:.3f}]")

# Define the diamond cut
cutg_diamond = ROOT.TCutG("cutg_diamond", 5)
cutg_diamond.SetVarX("Q2")
cutg_diamond.SetVarY("W")
cutg_diamond.SetPoint(0, vertex1[0], vertex1[1])  # bottom-left
cutg_diamond.SetPoint(1, vertex2[0], vertex2[1])  # top-left
cutg_diamond.SetPoint(2, vertex3[0], vertex3[1])  # top-right
cutg_diamond.SetPoint(3, vertex4[0], vertex4[1])  # bottom-right
cutg_diamond.SetPoint(4, vertex1[0], vertex1[1])  # bottom-left again to close the loop
Diamond_Cut = lambda event: (cutg_diamond.IsInside(event.Q2, event.W))

#------------------------------------------------------------------------------------------------------------------------------------

# Read the MMpi cut values from the CSV file
try:
    with open(mmcut_csv_file, mode='r', newline='') as csv_file:
        csv_reader = csv.DictReader(csv_file)
        for row in csv_reader:
            if row[csv_reader.fieldnames[0]].strip() == PHY_SETTING:  # Match first column (Physics_Setting)
                MM_Offset = float(row["MM_Offset"].strip())
                MM_Cut_lowvalue = float(row["MM_Cut_low"].strip())
                MM_Cut_highvalue = float(row["MM_Cut_high"].strip())
                break
        else:
            raise ValueError(f"No matching Physics_Setting '{PHY_SETTING}' found in {mmcut_csv_file}.")

except (FileNotFoundError, ValueError, KeyError) as e:
    print(f"Error: {e}")
    sys.exit(1)

# Print the assigned values
#print(f"MMpi_Offset = {MMpi_Offset:.6f}")
#print(f"MMpi_Cut_lowvalue = {MMpi_Cut_lowvalue}")
#print(f"MMpi_Cut_highvalue = {MMpi_Cut_highvalue}")
DATA_MMpi_Cut = lambda event: (MM_Cut_lowvalue <= (event.MMpi + MM_Offset) <= MM_Cut_highvalue)

#-----------------------------------------------------------------------------------------------------------------------------------------

# t-binning cut
# Read the CSV file
tbin_df = pd.read_csv(tbin_csv_file)

# Extract the t_min and t_max columns as arrays
t_min_values = tbin_df['t_min'].values
t_max_values = tbin_df['t_max'].values

# Define the cuts using row 1 and row 3 values
tbin_Cut1 = lambda event: (t_min_values[0] <= -event.MandelT <= t_max_values[0])  # Row 1
tbin_Cut2 = lambda event: (t_min_values[1] <= -event.MandelT <= t_max_values[1])  # Row 2
tbin_Cut3 = lambda event: (t_min_values[2] <= -event.MandelT <= t_max_values[2])  # Row 3
tbin_Cut4 = lambda event: (t_min_values[3] <= -event.MandelT <= t_max_values[3])  # Row 4
tbin_Cut5 = lambda event: (t_min_values[4] <= -event.MandelT <= t_max_values[4])  # Row 5

# Bundle them into a list
tbin_cuts = [tbin_Cut1, tbin_Cut2, tbin_Cut3, tbin_Cut4, tbin_Cut5]

# Print the t-binning cuts with 3 decimal places
print("\n t-binning cuts:")
print(f"tbin_Cut1: t_min = {t_min_values[0]:.3f}, t_max = {t_max_values[0]:.3f}")
print(f"tbin_Cut2: t_min = {t_min_values[1]:.3f}, t_max = {t_max_values[1]:.3f}")
print(f"tbin_Cut3: t_min = {t_min_values[2]:.3f}, t_max = {t_max_values[2]:.3f}")
print(f"tbin_Cut4: t_min = {t_min_values[3]:.3f}, t_max = {t_max_values[3]:.3f}")
print(f"tbin_Cut5: t_min = {t_min_values[4]:.3f}, t_max = {t_max_values[4]:.3f}")
print("-"*40)

# Define phi bins (15 bins from 0 to 360 degrees, each 24 degrees wide)
phi_bins = [i for i in range(0, 361, 24)]  # 0, 24, 48, ..., 360

#########################################################################################################################################

# Section for grabing Prompt/Random selection parameters from PARAM filePARAMPATH = UTILPATH+"/DB/PARAM"
# Function to read run numbers from a given run list file
def read_run_numbers(run_list_file):
    try:
        with open(run_list_file, 'r') as f:
            return [int(line.strip()) for line in f if line.strip().isdigit()]
    except FileNotFoundError:
        print(f"!!!!! ERROR !!!!!\nRun list file not found: {run_list_file}\n!!!!! ERROR !!!!!")
        sys.exit(2)

# Read run numbers from the specified files
data_run_numbers = read_run_numbers(data_run_list)
dummy_run_numbers = read_run_numbers(dummy_run_list)

# Section for grabbing Prompt/Random selection parameters from PARAM file
PARAMPATH = UTILPATH+"/DB/PARAM/PionLT"
TimingCutFile = "%s/PionLT_Timing_Parameters.csv" % PARAMPATH # This should match the param file actually being used!

TimingCutf = open(TimingCutFile)
try:
    TimingCutFile
except NameError:
    print("!!!!! ERRROR !!!!!\n One (or more) of the cut files not found!\n!!!!! ERRORR !!!!!")
    sys.exit(2)
print("\n Reading timing cuts from %s to select random windows \n" % TimingCutFile)

PromptWindow = [0, 0]
RandomWindows = [0, 0, 0, 0]
linenum = 0 # Count line number we're on
TempPar = -1 # To check later
for line in TimingCutf: # Read all lines in the cut file
    linenum += 1 # Add one to line number at start of loop
    if(linenum > 1): # Skip first line
        line = line.partition('#')[0] # Treat anything after a # as a comment and ignore it
        line = line.rstrip()
        array_list = line.split(",") # Convert line into an array_list, anything after a comma is a new entry

        # Check if any of the run numbers are within the specified range
        if any(run_num in range(int(array_list[0]), int(array_list[1]) + 1) for run_num in data_run_numbers + dummy_run_numbers):
            TempPar += 2  # If run number is in range, set to non -1 value
            BunchSpacing = float(array_list[2])
            CoinOffset = float(array_list[3])  # Coin offset value
            nSkip = float(array_list[4])  # Number of random windows skipped
            nWindows = float(array_list[5])  # Total number of random windows
            PromptPeak = float(array_list[6])  # Pion CT prompt peak position

if(TempPar == -1): # If value is still -1, run number provided din't match any ranges specified so exit
    print("!!!!! ERROR !!!!!\n Run number specified does not fall within a set of runs for which cuts are defined in %s\n!!!!! ERROR !!!!!" % TimingCutFile)
    sys.exit(3)
elif(TempPar > 1):
    print("!!! WARNING!!! Run number was found within the range of two (or more) line entries of %s !!! WARNING !!!" % TimingCutFile)
    print("The last matching entry will be treated as the input, you should ensure this is what you want")

# From our values from the file, reconstruct our windows
PromptWindow[0] = PromptPeak - (BunchSpacing/2) - CoinOffset
PromptWindow[1] = PromptPeak + (BunchSpacing/2) + CoinOffset
RandomWindows[0] = PromptPeak - (BunchSpacing/2) - CoinOffset - (nSkip*BunchSpacing) - ((nWindows/2)*BunchSpacing)
RandomWindows[1] = PromptPeak - (BunchSpacing/2) - CoinOffset - (nSkip*BunchSpacing)
RandomWindows[2] = PromptPeak + (BunchSpacing/2) + CoinOffset + (nSkip*BunchSpacing) 
RandomWindows[3] = PromptPeak + (BunchSpacing/2) + CoinOffset + (nSkip*BunchSpacing) + ((nWindows/2)*BunchSpacing)

###################################################################################################################################################

print("-"*40)
print ('Calculating effective charge for the data and dummy run list')
print("-"*40)

# Read CSV File and calculate total charge in mC for normalization
# Input runlists and csv files
data_run_list_file = (data_run_list)
dummy_run_list_file = (dummy_run_list)
eff_csv_file_name = (eff_csv_file)

# Read run numbers from the run list file
with open(data_run_list_file, 'r') as run_list_file_1:
    data_runs = [line.strip() for line in run_list_file_1 if line.strip()]
with open(dummy_run_list_file, 'r') as run_list_file_2:
    dummy_runs = [line.strip() for line in run_list_file_2 if line.strip()]

# Read CSV file using pandas
df = pd.read_csv(eff_csv_file_name)

# Filter DataFrame to include only rows with run numbers in the run list
filtered_data_df = df[df['Run_Number'].astype(str).str.replace('.0', '', regex=False).isin(data_runs)]
filtered_dummy_df = df[df['Run_Number'].astype(str).str.replace('.0', '', regex=False).isin(dummy_runs)]

# Intializing Variables to calculate the product of efficiency and charge
total_data_effective_charge = 0.0
total_dummy_effective_charge = 0.0
total_data_effective_charge_sum = 0.0
total_dummy_effective_charge_sum = 0.0
total_data_effective_charge_error_sum = 0.0
total_dummy_effective_charge_error_sum = 0.0
total_dummy_effective_charge_cal = 0.0
total_dummy_effective_charge_cal_error = 0.0

# Detector (HMS Cer + HMS Cal) Efficiencies
hms_Cer_detector_efficiency = 0.9981
hms_Cal_detector_efficiency = 0.9981
hms_Cer_detector_efficiency_error = 0.0001
hms_Cal_detector_efficiency_error = 0.0001

row_data = []

# Print charge values for each run
for index, row in filtered_data_df.iterrows():
    data_charge = row['BCM2_Beam_Cut_Charge']  # Assuming 'BCM2_Charge' is a column in your CSV
    data_hms_tracking_efficiency = row['HMS_Elec_SING_TRACK_EFF']  # Assuming 'HMS Tracking_Efficiency' is a column in your CSV
    data_shms_tracking_efficiency = row['SHMS_Prot_SING_TRACK_EFF']  # Assuming 'SHMS Tracking_Efficiency' is a column in your CSV
    data_hms_hodo_3_of_4_efficiency = row['HMS_Hodo_3_of_4_EFF']
    data_shms_hodo_3_of_4_efficiency = row['SHMS_Hodo_3_of_4_EFF']
#    data_shms_hodo_3_of_4_efficiency = 0.99
    data_edtm_livetime_Corr = row['Non_Scaler_EDTM_Live_Time_Corr']
    data_BCM2_Beam_Cut_Current = row['BCM2_Beam_Cut_Current']

    data_hms_tracking_efficiency_error = row['HMS_Elec_SING_TRACK_EFF_ERROR']
    data_shms_tracking_efficiency_error = row['SHMS_Prot_SING_TRACK_EFF_ERROR']
    data_edtm_livetime_Corr_error = row['Non_Scaler_EDTM_Live_Time_Corr_ERROR']
    data_hms_hodo_3_of_4_efficiency_error = row['HMS_Elec_SING_TRACK_EFF_ERROR']
    data_shms_hodo_3_of_4_efficiency_error = row['SHMS_Prot_SING_TRACK_EFF_ERROR']

#    data_hms_hodo_3_of_4_efficiency_error = row['HMS_Hodo_3_of_4_EFF_ERROR']
#    data_shms_hodo_3_of_4_efficiency_error = row['SHMS_Hodo_3_of_4_EFF_ERROR']
#    data_BCM2_Beam_Cut_Current_error = row['BCM2_Beam_Cut_Current_ERROR']

    # Nathan's, Richard's and Vijay's Boiling Correction
    data_Boiling_factor = 1 + (-0.00028 * data_BCM2_Beam_Cut_Current)
#    data_Boiling_factor = 1 + (-0.0007899 * data_BCM2_Beam_Cut_Current)
#    data_Boiling_factor = 1 + (-0.0005296 * data_BCM2_Beam_Cut_Current)
#    data_Boiling_factor_error = abs(-0.0007899 * data_BCM2_Beam_Cut_Current_error)

    data_BCM2_Beam_Cut_Current_error = 0.0
    data_Boiling_factor_error = 0.0

    # Calculating Current and Charge uncertainties
    data_run_number = row['Run_Number']
    data_run_time = row['BCM_Cut_HMS_Run_Length']
    if data_run_number <= 11988:
        data_slope = 5598.59 * 1000
        data_d_slope = 26.7 * 1000
        data_amplitude = 250356
        data_d_amplitude = 606.91
    elif 11988 < data_run_number <= 14777:
        data_slope = 5513.79 * 1000
        data_d_slope = 7.6  * 1000
        data_amplitude = 250429
        data_d_amplitude = 271
    elif 14777 < data_run_number <= 17000:
        data_slope = 5542 * 1000
        data_d_slope = 11.5  * 1000
        data_amplitude = 250029
        data_d_amplitude = 270
    else:
        print("Requires correct run number")
        continue

    data_charge_error = math.sqrt(((data_charge * data_d_slope)**2 + (data_d_amplitude * data_run_time)**2)/(data_slope)**2)
    data_current_error = math.sqrt((data_d_slope/data_slope)**2 + (data_d_amplitude / (data_slope * data_BCM2_Beam_Cut_Current))**2)
    data_Boiling_factor_error = math.sqrt((data_BCM2_Beam_Cut_Current_error/data_BCM2_Beam_Cut_Current)** 2 + (0.000017/0.00028)**2)

#    data_charge_error = data_charge * math.sqrt((d_slope/slope)**2 + ((d_amplitude * data_run_time)/(slope * data_charge))**2)
#    data_current_error = data_BCM2_Beam_Cut_Current * math.sqrt((d_slope/slope)**2 + (d_amplitude / (slope * data_BCM2_Beam_Cut_Current))**2)
#    data_Boiling_factor_error = data_Boiling_factor * math.sqrt((data_BCM2_Beam_Cut_Current_error/data_BCM2_Beam_Cut_Current)** 2 + (0.000017/0.00028)**2)

    data_product = (data_charge * data_hms_tracking_efficiency * data_shms_tracking_efficiency * hms_Cer_detector_efficiency * hms_Cal_detector_efficiency * data_hms_hodo_3_of_4_efficiency * data_shms_hodo_3_of_4_efficiency * data_edtm_livetime_Corr * data_Boiling_factor)
    data_product_error = data_product * (math.sqrt((data_charge_error/data_charge)** 2 +  (data_hms_tracking_efficiency_error/data_hms_tracking_efficiency)** 2 + (data_shms_tracking_efficiency_error/data_shms_tracking_efficiency)** 2 + (data_edtm_livetime_Corr_error/data_edtm_livetime_Corr)** 2 + (hms_Cer_detector_efficiency_error/hms_Cer_detector_efficiency)** 2 + (hms_Cal_detector_efficiency_error/hms_Cal_detector_efficiency)** 2 + (data_hms_hodo_3_of_4_efficiency_error/data_hms_hodo_3_of_4_efficiency)** 2 + (data_shms_hodo_3_of_4_efficiency_error/data_shms_hodo_3_of_4_efficiency)** 2 + (data_Boiling_factor_error/data_Boiling_factor)** 2)) 

    total_data_effective_charge_sum += data_product
    total_data_effective_charge_error_sum += (data_product_error)** 2

    print('Charge for data run: {:<10} Data Charge: {:.3f} HMS Tracking Eff: {:.3f} SHMS Tracking Eff: {:.3f} HMS Cer Detector Eff: {:.3f} HMS Cal Detector_Eff: {:.3f} HMS Hodo 3/4 Eff: {:.3f} SHMS Hodo 3/4 Eff: {:.3f} EDTM Live Time: {:.3f} Data Current: {:.3f} Boiling Correction: {:.3f} Product: {:.3f}'.format(row["Run_Number"], data_charge, data_hms_tracking_efficiency, data_shms_tracking_efficiency, hms_Cer_detector_efficiency, hms_Cal_detector_efficiency, data_hms_hodo_3_of_4_efficiency, data_shms_hodo_3_of_4_efficiency, data_edtm_livetime_Corr, data_BCM2_Beam_Cut_Current, data_Boiling_factor, data_product))

    row_data.append({'Run_Number':row["Run_Number"], 'charge':data_charge, 'charge_error':data_charge_error, 'HMS_Tracking_Eff':data_hms_tracking_efficiency, 'HMS_Tracking_Eff_error':data_hms_tracking_efficiency_error, 'SHMS_Tracking_Eff':data_shms_tracking_efficiency, 'SHMS_Tracking_Eff_error':data_shms_tracking_efficiency_error, 'HMS_Cer_Detector_Eff':hms_Cer_detector_efficiency, 'HMS_Cer_Detector_Eff_error':hms_Cer_detector_efficiency_error, 'HMS_Cal_Detector_Eff':hms_Cal_detector_efficiency, 'HMS_Cal_Detector_Eff_error':hms_Cal_detector_efficiency_error, 'HMS_Hodo_3_4_Eff':data_hms_hodo_3_of_4_efficiency, 'HMS_Hodo_3_4_Eff_error':data_hms_hodo_3_of_4_efficiency_error, 'SHMS_Hodo_3_4_Eff':data_shms_hodo_3_of_4_efficiency, 'SHMS_Hodo_3_4_Eff_error':data_shms_hodo_3_of_4_efficiency_error, 'EDTM_Live_Time':data_edtm_livetime_Corr, 'EDTM_Live_Time_error':data_edtm_livetime_Corr_error, 'Boiling_factor':data_Boiling_factor, 'Boiling_factor_error':data_Boiling_factor_error, 'effective_charge':data_product, 'effective_charge_error':data_product_error})

print("-"*40)

total_data_effective_charge = total_data_effective_charge_sum
total_data_effective_charge_error = math.sqrt(total_data_effective_charge_error_sum)

#------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

row_dummy = []

for index, row in filtered_dummy_df.iterrows():
    dummy_charge = row['BCM2_Beam_Cut_Charge']  # Assuming 'BCM2_Charge' is a column in your CSV
    dummy_hms_tracking_efficiency = row['HMS_Elec_SING_TRACK_EFF']  # Assuming 'HMS Efficiency' is a column in your CSV
    dummy_shms_tracking_efficiency = row['SHMS_Prot_SING_TRACK_EFF']  # Assuming 'SHMS Efficiency' is a column in your CSV
    dummy_hms_hodo_3_of_4_efficiency = row['HMS_Hodo_3_of_4_EFF']
    dummy_shms_hodo_3_of_4_efficiency = row['SHMS_Hodo_3_of_4_EFF']
#    dummy_shms_hodo_3_of_4_efficiency = 0.99
    dummy_edtm_livetime_Corr = row['Non_Scaler_EDTM_Live_Time_Corr']

#    dummy_charge_error = row['BCM2_Charge_ERROR']
    dummy_hms_tracking_efficiency_error = row['HMS_Elec_SING_TRACK_EFF_ERROR']
    dummy_shms_tracking_efficiency_error = row['SHMS_Prot_SING_TRACK_EFF_ERROR']
    dummy_edtm_livetime_Corr_error = row['Non_Scaler_EDTM_Live_Time_Corr_ERROR']
    dummy_hms_hodo_3_of_4_efficiency_error = row['HMS_Elec_SING_TRACK_EFF_ERROR']
    dummy_shms_hodo_3_of_4_efficiency_error = row['SHMS_Prot_SING_TRACK_EFF_ERROR']

#    dummy_hms_hodo_3_of_4_efficiency_error = row['HMS_Hodo_3_of_4_EFF_ERROR']
#    dummy_shms_hodo_3_of_4_efficiency_error = row['SHMS_Hodo_3_of_4_EFF_ERROR']

    # Calculating Current and Charge uncertainities
    dummy_run_number = row['Run_Number']
    dummy_run_time = row['BCM_Cut_HMS_Run_Length']
    if dummy_run_number <= 11988:
        dummy_slope = 5598.59 * 1000
        dummy_d_slope = 26.7 * 1000
        dummy_amplitude = 250356
        dummy_d_amplitude = 606.91
    elif 11988 < dummy_run_number <= 14777:
        dummy_slope = 5513.79 * 1000
        dummy_d_slope = 7.6 * 1000
        dummy_amplitude = 250429
        dummy_d_amplitude = 271
    elif 14777 < dummy_run_number <= 17000:
        dummy_slope = 5542 * 1000
        dummy_d_slope = 11.5 * 1000
        dummy_amplitude = 250029
        dummy_d_amplitude = 270
    else:
        print("Requires correct run number")
        continue

    dummy_charge_error = math.sqrt(((dummy_charge * dummy_d_slope)**2 + (dummy_d_amplitude * dummy_run_time)**2)/(dummy_slope)**2)

#    dummy_charge_error = dummy_charge * math.sqrt((d_slope/slope)**2 + ((d_amplitude * dummy_run_time)/(slope * dummy_charge))**2)
#    dummy_current_error = dummy_BCM2_Beam_Cut_Current * math.sqrt((d_slope/slope)**2 + (d_amplitude / (slope * dummy_BCM2_Beam_Cut_Current))**2)

    dummy_product = (dummy_charge * dummy_hms_tracking_efficiency * dummy_shms_tracking_efficiency * hms_Cer_detector_efficiency * hms_Cal_detector_efficiency * dummy_hms_hodo_3_of_4_efficiency * dummy_shms_hodo_3_of_4_efficiency * dummy_edtm_livetime_Corr)
    dummy_product_error = dummy_product * (math.sqrt(((dummy_charge_error/dummy_charge) ** 2 + dummy_hms_tracking_efficiency_error/dummy_hms_tracking_efficiency) ** 2 + (dummy_shms_tracking_efficiency_error/dummy_shms_tracking_efficiency) ** 2 + (dummy_edtm_livetime_Corr_error/dummy_edtm_livetime_Corr)** 2 + (hms_Cer_detector_efficiency_error/hms_Cer_detector_efficiency) ** 2 + (hms_Cal_detector_efficiency_error/hms_Cal_detector_efficiency) ** 2 + (dummy_hms_hodo_3_of_4_efficiency_error/dummy_hms_hodo_3_of_4_efficiency) ** 2 + (dummy_shms_hodo_3_of_4_efficiency_error/dummy_shms_hodo_3_of_4_efficiency) ** 2))

    total_dummy_effective_charge_sum += dummy_product
    total_dummy_effective_charge_error_sum += (dummy_product_error)** 2

    print('Charge for dummy run: {:<10} Dummy Charge: {:.3f} HMS Tracking Eff: {:.3f} SHMS Tracking Eff: {:.3f} HMS Cer Detector Eff: {:.3f} HMS Cal Detector_Eff: {:.3f} HMS Hodo 3/4 Eff: {:.3f} SHMS Hodo 3/4 Eff: {:.3f} EDTM Live Time: {:.3f} Product: {:.3f}'.format(row["Run_Number"], dummy_charge, dummy_hms_tracking_efficiency, dummy_shms_tracking_efficiency, hms_Cer_detector_efficiency, hms_Cal_detector_efficiency, dummy_hms_hodo_3_of_4_efficiency, dummy_shms_hodo_3_of_4_efficiency, dummy_edtm_livetime_Corr, dummy_product))

    row_dummy.append({'Run_Number':row["Run_Number"], 'charge':dummy_charge, 'charge_error':dummy_charge_error, 'HMS_Tracking_Eff':dummy_hms_tracking_efficiency, 'HMS_Tracking_Eff_error':dummy_hms_tracking_efficiency_error, 'SHMS_Tracking_Eff':dummy_shms_tracking_efficiency, 'SHMS_Tracking_Eff_error':dummy_shms_tracking_efficiency_error, 'HMS_Cer_Detector_Eff':hms_Cer_detector_efficiency, 'HMS_Cer_Detector_Eff_error':hms_Cer_detector_efficiency_error, 'HMS_Cal_Detector_Eff':hms_Cal_detector_efficiency, 'HMS_Cal_Detector_Eff_error':hms_Cal_detector_efficiency_error, 'HMS_Hodo_3_4_Eff':dummy_hms_hodo_3_of_4_efficiency, 'HMS_Hodo_3_4_Eff_error':dummy_hms_hodo_3_of_4_efficiency_error, 'SHMS_Hodo_3_4_Eff':dummy_shms_hodo_3_of_4_efficiency, 'SHMS_Hodo_3_4_Eff_error':dummy_shms_hodo_3_of_4_efficiency_error, 'EDTM_Live_Time':dummy_edtm_livetime_Corr, 'EDTM_Live_Time_error':dummy_edtm_livetime_Corr_error, 'Boiling_factor':'NA', 'Boiling_factor_error':'NA', 'effective_charge':dummy_product, 'effective_charge_error':dummy_product_error})

print("-"*40)

# Dummy Target Thickness Correction
#dummy_target_corr = 4.8579 # KaonLT
dummy_target_corr = 3.527 # PionLT
dummy_target_corr_error = 0.227 # PionLT

total_dummy_effective_charge_cal = total_dummy_effective_charge_sum
total_dummy_effective_charge_cal_error = math.sqrt(total_dummy_effective_charge_error_sum)

total_dummy_effective_charge = (total_dummy_effective_charge_cal * dummy_target_corr)
total_dummy_effective_charge_error = total_dummy_effective_charge * math.sqrt((total_dummy_effective_charge_cal_error/total_dummy_effective_charge_cal)**2 + (dummy_target_corr_error/dummy_target_corr)** 2)

#total_dummy_effective_charge = total_dummy_effective_charge_sum * dummy_target_corr
#total_dummy_effective_charge_error = math.sqrt(total_dummy_effective_charge_error_sum + (dummy_target_corr_error/dummy_target_corr)** 2)

print("\nTotal effective charge for the data run list: {:.5f} ± {:.5f}".format(total_data_effective_charge, total_data_effective_charge_error))
print("\nTotal effective charge for the dummy run list: {:.5f} ± {:.5f}".format(total_dummy_effective_charge, total_dummy_effective_charge_error))

# Normalization factor Calculation for Data and Dummy
normfac_data = 1.0/(total_data_effective_charge)
normfac_dummy = 1.0/(total_dummy_effective_charge)

print("-"*40)
print ("normfac_data :", normfac_data)
print ("normfac_dummy: ", normfac_dummy)
print("-"*40)

#------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

# Define the output CSV file name
eff_cal_output_csv_path = "%s/LTSep_CSVs/physics_yields_csv/%s_Physics_Norm_Eff_Charge_Calculation.csv" % (UTILPATH, setting_name)

# Prepare the header for the CSV file
header = [
    "Physics Setting", "Run Type", "Run Number", "Charge", "Charge Error",
    "HMS Tracking Eff", "HMS Tracking Eff Error", "SHMS Tracking Eff", "SHMS Tracking Eff Error",
    "HMS Cer Detector Eff", "HMS Cer Detector Eff Error", "HMS Cal Detector Eff", "HMS Cal Detector Eff Error",
    "HMS Hodo 3/4 Eff", "HMS Hodo 3/4 Eff Error", "SHMS Hodo 3/4 Eff", "SHMS Hodo 3/4 Eff Error",
    "EDTM Live Time", "EDTM Live Time Error", "Boiling Factor", "Boiling Factor Error", "Dummy Target Correction",
    "Dummy Target Correction Error", "Effective Charge", "Effective Charge Error", "Total Effective Charge", "Total Effective Charge Error", "Normfactor"
]

# Combine data and dummy rows
combined_rows = []

# Add data rows
for row in row_data:
    combined_rows.append({
        "Physics Setting": PHY_SETTING,
        "Run Type": "data",
        "Run Number": row["Run_Number"],
        "Charge": row["charge"],
        "Charge Error": row["charge_error"],
        "HMS Tracking Eff": row["HMS_Tracking_Eff"],
        "HMS Tracking Eff Error": row["HMS_Tracking_Eff_error"],
        "SHMS Tracking Eff": row["SHMS_Tracking_Eff"],
        "SHMS Tracking Eff Error": row["SHMS_Tracking_Eff_error"],
        "HMS Cer Detector Eff": row["HMS_Cer_Detector_Eff"],
        "HMS Cer Detector Eff Error": row["HMS_Cer_Detector_Eff_error"],
        "HMS Cal Detector Eff": row["HMS_Cal_Detector_Eff"],
        "HMS Cal Detector Eff Error": row["HMS_Cal_Detector_Eff_error"],
        "HMS Hodo 3/4 Eff": row["HMS_Hodo_3_4_Eff"],
        "HMS Hodo 3/4 Eff Error": row["HMS_Hodo_3_4_Eff_error"],
        "SHMS Hodo 3/4 Eff": row["SHMS_Hodo_3_4_Eff"],
        "SHMS Hodo 3/4 Eff Error": row["SHMS_Hodo_3_4_Eff_error"],
        "EDTM Live Time": row["EDTM_Live_Time"],
        "EDTM Live Time Error": row["EDTM_Live_Time_error"],
        "Boiling Factor": row["Boiling_factor"],
        "Boiling Factor Error": row["Boiling_factor_error"],
        "Dummy Target Correction": "NA",
        "Dummy Target Correction Error": "NA",
        "Effective Charge": row["effective_charge"],
        "Effective Charge Error": row["effective_charge_error"],
        "Total Effective Charge": total_data_effective_charge,
        "Total Effective Charge Error": total_data_effective_charge_error,
        "Normfactor": normfac_data
    })

# Add dummy rows
for row in row_dummy:
    combined_rows.append({
        "Physics Setting": PHY_SETTING,
        "Run Type": "dummy",
        "Run Number": row["Run_Number"],
        "Charge": row["charge"],
        "Charge Error": row["charge_error"],
        "HMS Tracking Eff": row["HMS_Tracking_Eff"],
        "HMS Tracking Eff Error": row["HMS_Tracking_Eff_error"],
        "SHMS Tracking Eff": row["SHMS_Tracking_Eff"],
        "SHMS Tracking Eff Error": row["SHMS_Tracking_Eff_error"],
        "HMS Cer Detector Eff": row["HMS_Cer_Detector_Eff"],
        "HMS Cer Detector Eff Error": row["HMS_Cer_Detector_Eff_error"],
        "HMS Cal Detector Eff": row["HMS_Cal_Detector_Eff"],
        "HMS Cal Detector Eff Error": row["HMS_Cal_Detector_Eff_error"],
        "HMS Hodo 3/4 Eff": row["HMS_Hodo_3_4_Eff"],
        "HMS Hodo 3/4 Eff Error": row["HMS_Hodo_3_4_Eff_error"],
        "SHMS Hodo 3/4 Eff": row["SHMS_Hodo_3_4_Eff"],
        "SHMS Hodo 3/4 Eff Error": row["SHMS_Hodo_3_4_Eff_error"],
        "EDTM Live Time": row["EDTM_Live_Time"],
        "EDTM Live Time Error": row["EDTM_Live_Time_error"],
        "Boiling Factor": row["Boiling_factor"],
        "Boiling Factor Error": row["Boiling_factor_error"],
        "Dummy Target Correction": dummy_target_corr,
        "Dummy Target Correction Error": dummy_target_corr_error,
        "Effective Charge": row["effective_charge"],
        "Effective Charge Error": row["effective_charge_error"],
        "Total Effective Charge": total_dummy_effective_charge,
        "Total Effective Charge Error": total_dummy_effective_charge_error,
        "Normfactor": normfac_dummy
    })

# Read the existing CSV file into memory
existing_rows = []
try:
    with open(eff_cal_output_csv_path, mode='r') as csv_file:
        csv_reader = csv.DictReader(csv_file)
        existing_rows = list(csv_reader)
except FileNotFoundError:
    # If the file doesn't exist, initialize an empty list
    existing_rows = []

# Update or add rows
updated_rows = []
for new_row in combined_rows:
    row_found = False
    for existing_row in existing_rows:
        if existing_row["Physics Setting"] == new_row["Physics Setting"] and existing_row["Run Type"] == new_row["Run Type"] and existing_row["Run Number"] == str(new_row["Run Number"]):
            # Update the existing row
            existing_row.update(new_row)
            row_found = True
            break
    if not row_found:
        # Add the new row if it doesn't exist
        updated_rows.append(new_row)

# Combine updated rows with existing rows
existing_rows.extend(updated_rows)

# Write everything back to the CSV file
with open(eff_cal_output_csv_path, mode='w', newline='') as csv_file:
    writer = csv.DictWriter(csv_file, fieldnames=header)
    writer.writeheader()
    writer.writerows(existing_rows)

print(f"Run factors written to {eff_cal_output_csv_path}")

###################################################################################################################################################

nbins = 200

hist_config = {
    "MMpi":     ("MIssing Mass Distribution (Cut_All); MM_{\pi}; Counts", nbins, 0.8, 1.2),
}

def initialize_histograms(t_min_values, t_max_values, phi_bins, hist_config, category_suffixes):
    histograms_by_tphi_cut = []
    for i in range(len(t_min_values)):
        tmin = t_min_values[i]
        tmax = t_max_values[i]
        t_label = f"t({tmin:.2f}-{tmax:.2f})".replace('.', 'p')
        for j in range(len(phi_bins) - 1):
            phimin = phi_bins[j]
            phimax = phi_bins[j + 1]
            phi_label = f"phi({phimin}-{phimax})"
            hist_set = {suffix: {} for suffix in category_suffixes}
            for var, (base_title, nbins, xmin, xmax) in hist_config.items():
                for suffix in category_suffixes:
                    # Create unique histogram name and title for each category
                    hname = f"{var}_{t_label}_{phi_label}_{suffix}_data_cut_all"
                    title = f"{base_title} (t ∈ [{tmin:.2f}, {tmax:.2f}], φ ∈ [{phimin}, {phimax}], Category: {suffix})"
                    hist_set[suffix][var] = ROOT.TH1D(hname, title, nbins, xmin, xmax)
            histograms_by_tphi_cut.append((tmin, tmax, phimin, phimax, hist_set))
    return histograms_by_tphi_cut

categories_data = ["prompt_data", "random_data", "randsub_data", "scaled_randsub_data", "prompt_dummy", "random_dummy", "randsub_dummy", "scaled_randsub_dummy", "dummysub"]

# Initialize histograms for data
data_histograms_by_tphi_cut = initialize_histograms(t_min_values, t_max_values, phi_bins, hist_config, categories_data)

##########################################################################################################################################################################################################

# Read stuff from the main event tree
infile_DATA = ROOT.TFile.Open(rootFile_DATA, "READ")
infile_DUMMY = ROOT.TFile.Open(rootFile_DUMMY, "READ")

#Uncut_Pion_Events_Data_tree = infile_DATA.Get("Uncut_Pion_Events")
#Cut_Pion_Events_Accpt_Data_tree = infile_DATA.Get("Cut_Pion_Events_Accpt")
#Cut_Pion_Events_All_Data_tree = infile_DATA.Get("Cut_Pion_Events_All")
Cut_Pion_Events_Prompt_Data_tree = infile_DATA.Get("Cut_Pion_Events_Prompt")
Cut_Pion_Events_Random_Data_tree = infile_DATA.Get("Cut_Pion_Events_Random")
#nEntries_TBRANCH_DATA  = Cut_Pion_Events_Prompt_Data_tree.GetEntries()

#Uncut_Pion_Events_Dummy_tree = infile_DUMMY.Get("Uncut_Pion_Events")
#Cut_Pion_Events_Accpt_Dummy_tree = infile_DUMMY.Get("Cut_Pion_Events_Accpt")
#Cut_Pion_Events_All_Dummy_tree = infile_DUMMY.Get("Cut_Pion_Events_All")
Cut_Pion_Events_Prompt_Dummy_tree = infile_DUMMY.Get("Cut_Pion_Events_Prompt")
Cut_Pion_Events_Random_Dummy_tree = infile_DUMMY.Get("Cut_Pion_Events_Random")
#nEntries_TBRANCH_DUMMY  = Cut_Pion_Events_Prompt_Dummy_tree.GetEntries()

###################################################################################################################################################

# Fill the histograms with the data
for event in Cut_Pion_Events_Prompt_Data_tree:
    # Apply the MMpi cut and Diamond cut
    if DATA_MMpi_Cut(event) and Diamond_Cut(event):
        # Loop over t-bins
        for tbin_cut, (tmin, tmax) in zip(tbin_cuts, zip(t_min_values, t_max_values)):
            if tbin_cut(event):  # Apply the t-bin cut
                # Loop over phi-bins
                for phimin, phimax in zip(phi_bins[:-1], phi_bins[1:]):
                    if phimin <= event.ph_q * (180 / math.pi) + 180 <= phimax:  # Apply the phi-bin cut
                        # Fill the histogram for the corresponding t-bin and phi-bin
                        for tmin_hist, tmax_hist, phimin_hist, phimax_hist, hist_set in data_histograms_by_tphi_cut:
                            if tmin == tmin_hist and tmax == tmax_hist and phimin == phimin_hist and phimax == phimax_hist:
                                for var in hist_config:
                                    hist_set["prompt_data"][var].Fill(event.MMpi + MM_Offset)

for event in Cut_Pion_Events_Random_Data_tree:
    # Apply the MMpi cut and Diamond cut
    if DATA_MMpi_Cut(event) and Diamond_Cut(event):
        # Loop over t-bins
        for tbin_cut, (tmin, tmax) in zip(tbin_cuts, zip(t_min_values, t_max_values)):
            if tbin_cut(event):  # Apply the t-bin cut
                # Loop over phi-bins
                for phimin, phimax in zip(phi_bins[:-1], phi_bins[1:]):
                    if phimin <= event.ph_q * (180 / math.pi) + 180 <= phimax:  # Apply the phi-bin cut
                        # Fill the histogram for the corresponding t-bin and phi-bin
                        for tmin_hist, tmax_hist, phimin_hist, phimax_hist, hist_set in data_histograms_by_tphi_cut:
                            if tmin == tmin_hist and tmax == tmax_hist and phimin == phimin_hist and phimax == phimax_hist:
                                for var in hist_config:
                                    hist_set["random_data"][var].Fill(event.MMpi + MM_Offset)

for event in Cut_Pion_Events_Prompt_Dummy_tree:
    # Apply the MMpi cut and Diamond cut
    if DATA_MMpi_Cut(event) and Diamond_Cut(event):
        # Loop over t-bins
        for tbin_cut, (tmin, tmax) in zip(tbin_cuts, zip(t_min_values, t_max_values)):
            if tbin_cut(event):  # Apply the t-bin cut
                # Loop over phi-bins
                for phimin, phimax in zip(phi_bins[:-1], phi_bins[1:]):
                    if phimin <= event.ph_q * (180 / math.pi) + 180 <= phimax:  # Apply the phi-bin cut
                        # Fill the histogram for the corresponding t-bin and phi-bin
                        for tmin_hist, tmax_hist, phimin_hist, phimax_hist, hist_set in data_histograms_by_tphi_cut:
                            if tmin == tmin_hist and tmax == tmax_hist and phimin == phimin_hist and phimax == phimax_hist:
                                for var in hist_config:
                                    hist_set["prompt_dummy"][var].Fill(event.MMpi + MM_Offset)

for event in Cut_Pion_Events_Random_Dummy_tree:
    # Apply the MMpi cut and Diamond cut
    if DATA_MMpi_Cut(event) and Diamond_Cut(event):
        # Loop over t-bins
        for tbin_cut, (tmin, tmax) in zip(tbin_cuts, zip(t_min_values, t_max_values)):
            if tbin_cut(event):  # Apply the t-bin cut
                # Loop over phi-bins
                for phimin, phimax in zip(phi_bins[:-1], phi_bins[1:]):
                    if phimin <= event.ph_q * (180 / math.pi) + 180 <= phimax:  # Apply the phi-bin cut
                        # Fill the histogram for the corresponding t-bin and phi-bin
                        for tmin_hist, tmax_hist, phimin_hist, phimax_hist, hist_set in data_histograms_by_tphi_cut:
                            if tmin == tmin_hist and tmax == tmax_hist and phimin == phimin_hist and phimax == phimax_hist:
                                for var in hist_config:
                                    hist_set["random_dummy"][var].Fill(event.MMpi + MM_Offset)

print("####################################")
print("###### Histogram filling done ######")
print("####################################\n")

#################################################################################################################################################

# Scale random histograms
for tmin_hist, tmax_hist, phimin_hist, phimax_hist, hist_set in data_histograms_by_tphi_cut:
    for var in hist_config:
        # Scale the random data histogram
        random_data_hist = hist_set["random_data"][var]
        if random_data_hist:
            random_data_hist.Scale(1/nWindows)
            # Save the scaled histogram back into the random_data category
            hist_set["random_data"][var] = random_data_hist

        # Scale the random dummy histogram
        random_dummy_hist = hist_set["random_dummy"][var]
        if random_dummy_hist:
            random_dummy_hist.Scale(1/nWindows)
            # Save the scaled histogram back into the random_data category
            hist_set["random_dummy"][var] = random_dummy_hist

# Perform random subtraction for each t and phi bin
for tmin_hist, tmax_hist, phimin_hist, phimax_hist, hist_set in data_histograms_by_tphi_cut:
    for var in hist_config:
        # Get the prompt and random histograms
        prompt_data_hist = hist_set["prompt_data"][var]
        random_data_hist = hist_set["random_data"][var]
        # Directly subtract the random histogram from the prompt histogram
        randsub_data_hist = prompt_data_hist.Clone()
        randsub_data_hist.Add(random_data_hist, -1)
        # Store the result in the "randsub_data" category
        hist_set["randsub_data"][var] = randsub_data_hist

for tmin_hist, tmax_hist, phimin_hist, phimax_hist, hist_set in data_histograms_by_tphi_cut:
    for var in hist_config:
        # Get the prompt and random histograms for dummy
        prompt_dummy_hist = hist_set["prompt_dummy"][var]
        random_dummy_hist = hist_set["random_dummy"][var]
        # Directly subtract the random dummy histogram from the prompt dummy histogram
        randsub_dummy_hist = prompt_dummy_hist.Clone()
        randsub_dummy_hist.Add(random_dummy_hist, -1)
        # Store the result in the "randsub_dummy" category
        hist_set["randsub_dummy"][var] = randsub_dummy_hist

print("######################################")
print("###### Random substraction done ######")
print("######################################\n")

############################################################################################################################################

# Scale randsub_data and randsub_dummy histograms with normalization factors
for tmin_hist, tmax_hist, phimin_hist, phimax_hist, hist_set in data_histograms_by_tphi_cut:
    for var in hist_config:
        # Scale the randsub_data histogram
        randsub_data_hist = hist_set["randsub_data"][var]
        scaled_randsub_data_hist = randsub_data_hist.Clone()
        scaled_randsub_data_hist.Scale(normfac_data)
        hist_set["scaled_randsub_data"][var] = scaled_randsub_data_hist

        # Scale the randsub_dummy histogram
        randsub_dummy_hist = hist_set["randsub_dummy"][var]
        scaled_randsub_dummy_hist = randsub_dummy_hist.Clone()
        scaled_randsub_dummy_hist.Scale(normfac_dummy)
        hist_set["scaled_randsub_dummy"][var] = scaled_randsub_dummy_hist

# Dummy Subtraction
for tmin_hist, tmax_hist, phimin_hist, phimax_hist, hist_set in data_histograms_by_tphi_cut:
    for var in hist_config:
        # Get the randsub_data and randsub_dummy histograms
        scaled_randsub_data_hist = hist_set["scaled_randsub_data"][var]
        scaled_randsub_dummy_hist = hist_set["scaled_randsub_dummy"][var]
        # Directly subtract the randsub_dummy histogram from the randsub_data histogram
        dummysub_data_hist = scaled_randsub_data_hist.Clone()
        dummysub_data_hist.Add(scaled_randsub_dummy_hist, -1)
        # Store the result in the "dummysub" category
        hist_set["dummysub"][var] = dummysub_data_hist

        # Print confirmation of subtraction
        #print(f"Subtraction complete for t-bin: [{tmin_hist}, {tmax_hist}], phi-bin: [{phimin_hist}, {phimax_hist}], variable: {var}")

print("####################################")
print("###### Dummy subtraction done ######")
print("####################################\n")

#############################################################################################################################################

# Physics Yield Calculations
# Dictionary to store integrals and errors for each t-bin and phi-bin
dN_data_MMpi = {}

# Loop over all t-bins and phi-bins
for tmin, tmax, phimin, phimax, hist_set in data_histograms_by_tphi_cut:
    # Create a unique key for the current t-bin and phi-bin
    bin_key = f"t({tmin:.2f}-{tmax:.2f})_phi({phimin}-{phimax})"

    # Initialize the error array
    dN_data = array.array('d', [0.0])

    # Get the dummysub histogram for the current bin
    hist = hist_set["dummysub"]["MMpi"]  # Replace "MMpi" with the variable of interest

    # Calculate the integral and error
    N_data = hist.IntegralAndError(1, nbins, dN_data, "")

    # Store the results in the dictionary
    dN_data_MMpi[bin_key] = {
        "integral": N_data,
        "error": dN_data[0]
    }

    # Print the result for debugging
#    print(f"Physics Yield: {bin_key} = Yield: {N_data:.8f}, Error: {dN_data[0]:.8f}")
#print("-" * 40)

# Dictionary to store integrals and errors for each t-bin and phi-bin
dN_error = {}

# Loop over all t-bins and phi-bins
for tmin, tmax, phimin, phimax, hist_set in data_histograms_by_tphi_cut:
    # Create a unique key for the current t-bin and phi-bin
    bin_key = f"t({tmin:.2f}-{tmax:.2f})_phi({phimin}-{phimax})"

    # Initialize error arrays
    dN_data_error = array.array('d', [0.0])
    dN_dummy_error = array.array('d', [0.0])

    # Get the randsub_data and randsub_dummy histograms for the current bin
    randsub_data_hist = hist_set["randsub_data"]["MMpi"]  # Replace "MMpi" with the variable of interest
    randsub_dummy_hist = hist_set["randsub_dummy"]["MMpi"]  # Replace "MMpi" with the variable of interest

    # Calculate the integrals and errors
    N_data = randsub_data_hist.IntegralAndError(1, nbins, dN_data_error, "")
    N_dummy = randsub_dummy_hist.IntegralAndError(1, nbins, dN_dummy_error, "")
    N_dummysub = N_data - N_dummy

    # Normalize the integrals
    N_data_norm = N_data / total_data_effective_charge
    N_dummy_norm = N_dummy / total_dummy_effective_charge

    # Calculate the normalized errors
    if N_data != 0:
        dN_data_norm = N_data_norm * math.sqrt((dN_data_error[0] / N_data) ** 2 + (total_data_effective_charge_error / total_data_effective_charge) ** 2)
    else:
        dN_data_norm = N_data_norm * math.sqrt((0) ** 2 + (total_data_effective_charge_error / total_data_effective_charge) ** 2)
    if N_dummy != 0:
        dN_dummy_norm = N_dummy_norm * math.sqrt((dN_dummy_error[0] / N_dummy) ** 2 + (total_dummy_effective_charge_error / total_dummy_effective_charge) ** 2)
    else:
        dN_dummy_norm = N_data_norm * math.sqrt((0) ** 2 + (total_dummy_effective_charge_error / total_dummy_effective_charge) ** 2)

    # Ensure dN_data_norm and dN_dummy_norm are not zero
    if dN_data_norm == 0:
        dN_data_norm = 1.0

#    if dN_dummy_norm == 0:
#        dN_dummy_norm = 1.0

    # Calculate the final yield and its error
    N_data_dummy_sub_norm = N_data_norm - N_dummy_norm
    dN_data_dummy_sub_norm = math.sqrt(dN_data_norm ** 2 + dN_dummy_norm ** 2)

    # Store the results in the dictionary
    dN_error[bin_key] = {
        "Data Counts": N_data,
        "Dummy Counts": N_dummy,
        "N_dummysub": N_dummysub,
        "N_data_norm": N_data_norm,
        "N_dummy_norm": N_dummy_norm,
        "dN_data_norm": dN_data_norm,
        "dN_dummy_norm": dN_dummy_norm,
        "N_data_dummy_sub_norm": N_data_dummy_sub_norm,
        "dN_data_dummy_sub_norm": dN_data_dummy_sub_norm
    }

    # Print the result for debugging
    print(f"Physics Yield: {bin_key} = Yield: {N_data_dummy_sub_norm:.8f}, Error: {dN_data_dummy_sub_norm:.8f}")
print("-" * 40)

# Print the results from both loops
print("=" * 40)
print("Physics Yield Results:")
for bin_key in dN_data_MMpi:
    # Use the correct key names from dN_data_MMP
    Y_data = dN_data_MMpi[bin_key]["integral"]
    dY_data_dummy_sub_norm = dN_error[bin_key]["dN_data_dummy_sub_norm"]
#    print(f"Physics Yield: {bin_key} = Yield: {Y_data:.8f}, Error: {dY_data_dummy_sub_norm:.8f}")
#print("=" * 40)

#------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

# Define the output CSV file name
yields_pions_path =  "%s/LTSep_CSVs/physics_yields_csv/%s_Physics_Data_Yield.csv" % (UTILPATH, setting_name)

# Read the existing CSV file into memory
rows = []
header = ["Physics Setting", "t_min", "t_max", "phi_min", "phi_max", "DummySub_Counts", "physics_yield", "physics_yield_error", "%error/yield"]
try:
    with open(yields_pions_path, mode='r') as csv_file:
        csv_reader = csv.DictReader(csv_file)
        rows = list(csv_reader)
except FileNotFoundError:
    # If the file doesn't exist, initialize an empty list
    rows = []

# Update or add rows
updated_rows = []
for tmin, tmax, phimin, phimax, hist_set in data_histograms_by_tphi_cut:
    # Create a unique key for the current t-bin and phi-bin
    bin_key = f"t({tmin:.2f}-{tmax:.2f})_phi({phimin}-{phimax})"

    # Retrieve the physics yield (N_data) and its error
    N_dummysub = dN_error[bin_key]["N_dummysub"]
    Y_data = dN_data_MMpi[bin_key]["integral"]
    dY_data_dummy_sub_norm = dN_error[bin_key]["dN_data_dummy_sub_norm"]

    # Check if the setting already exists in the CSV
    row_found = False
    for row in rows:
        if (
            row["Physics Setting"] == PHY_SETTING
            and float(row["t_min"]) == tmin
            and float(row["t_max"]) == tmax
            and float(row["phi_min"]) == phimin
            and float(row["phi_max"]) == phimax
        ):
            # Update the existing row
            row.update({
                "Physics Setting": PHY_SETTING,
                "t_min": tmin,
                "t_max": tmax,
                "phi_min": phimin,
                "phi_max": phimax,
                "DummySub_Counts": N_dummysub,
                "physics_yield": Y_data,
                "physics_yield_error": dY_data_dummy_sub_norm,
                "%error/yield": (dY_data_dummy_sub_norm/Y_data)*100 if Y_data != 0 else dY_data_dummy_sub_norm
            })
            row_found = True
            break

    if not row_found:
        # Add a new row if the setting doesn't exist
        updated_rows.append({
            "Physics Setting": PHY_SETTING,
            "t_min": tmin,
            "t_max": tmax,
            "phi_min": phimin,
            "phi_max": phimax,
            "DummySub_Counts": N_dummysub,
            "physics_yield": Y_data,
            "physics_yield_error": dY_data_dummy_sub_norm,
            "%error/yield": (dY_data_dummy_sub_norm/Y_data)*100 if Y_data != 0 else dY_data_dummy_sub_norm

        })

# Combine updated rows with existing rows
rows.extend(updated_rows)

# Write everything back to the CSV
with open(yields_pions_path, mode='w', newline='') as csv_file:
    writer = csv.DictWriter(csv_file, fieldnames=header)
    writer.writeheader()
    writer.writerows(rows)

print(f"Physics yield results written to {yields_pions_path}")

'''
from openpyxl import Workbook
from openpyxl.styles import PatternFill

# Define the output Excel file name
output_excel_path = "%s/LTSep_CSVs/physics_yields_csv/%s_Physics_Data_Yield.xlsx" % (UTILPATH, setting_name)

# Create a new workbook and select the active worksheet
wb = Workbook()
ws = wb.active
ws.title = "Physics Yields"

# Define the light blue fill style
light_blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")

# Write the header
header = ["Physics Setting", "t_min", "t_max", "phi_min", "phi_max", "physics_yield", "physics_yield_error"]
ws.append(header)

# Apply the light blue fill to the header row
for cell in ws[1]:
    cell.fill = light_blue_fill

# Write the data rows
for row in rows:
    ws.append([
        row["Physics Setting"],
        row["t_min"],
        row["t_max"],
        row["phi_min"],
        row["phi_max"],
        row["physics_yield"],
        row["physics_yield_error"]
    ])

    # Apply the light blue fill to the current row
    for cell in ws[ws.max_row]:
        cell.fill = light_blue_fill

# Save the workbook
wb.save(output_excel_path)
'''
#############################################################################################################################################

# Making directories in output file
outHistFile = ROOT.TFile.Open("%s/%s_%s_Yield_Data.root" % (OUTPATH, PHY_SETTING, MaxEvent) , "RECREATE")

# Write histograms to t-bin-specific directories
for tmin, tmax, phimin, phimax, hist_set in data_histograms_by_tphi_cut:
    # Format the directory name for the t-bin
    t_dirname = f"Cut_Pion_Events_Norm_DummySub_t{tmin:.2f}-t{tmax:.2f}_Data".replace(".", "p")
    
    # Check if the directory already exists
    tdir = outHistFile.GetDirectory(t_dirname)
    if not tdir:
        # Create the directory if it doesn't exist
        tdir = outHistFile.mkdir(t_dirname)
    tdir.cd()

    # Write each phi-bin histogram for this t-bin
    for var in hist_config:
        for cat in ["dummysub"]:
            hist = hist_set[cat].get(var)
            if hist:
                # Update the histogram title to include t and phi ranges
                hist.SetTitle(f"{var}_t{tmin:.2f}-t{tmax:.2f}_phi{phimin}-phi{phimax}_norm_{cat}_data_cut_all".replace(".", "p"))
            
                # Update the Y-axis label to "Counts"
                hist.GetYaxis().SetTitle("Counts")

                # Set the histogram name dynamically
                hist.SetName(f"{var}_t{tmin:.2f}-t{tmax:.2f}_phi{phimin}-phi{phimax}_norm_{cat}_data_cut_all".replace(".", "p"))

                # Set the default drawing option to "HIST"
                #hist.SetOption("HIST")

                # Write the histogram to the file
                hist.Write()

print("####################################")
print("###### Histogram writing done ######")
print("####################################\n")

infile_DATA.Close() 
infile_DUMMY.Close()
outHistFile.Close()

print ("Processing Complete")